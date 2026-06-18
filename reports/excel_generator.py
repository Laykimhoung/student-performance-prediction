from openpyxl import Workbook
from tkinter import filedialog
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_class_excel(class_data, students):

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Performance" 

    # ==================================
    # STYLES
    # ==================================
    dark_blue = "1E3A8A"

    low_risk = "DCFCE7"
    medium_risk = "FEF3C7"
    high_risk = "FEE2E2"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=dark_blue
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    bold_font = Font(
        bold=True
    )

    # ==================================
    # RISK COUNTS
    # ==================================
    high_count = sum(
        1 for s in students
        if s["risk"] == "High"
    )

    medium_count = sum(
        1 for s in students
        if s["risk"] == "Medium"
    )

    low_count = sum(
        1 for s in students
        if s["risk"] == "Low"
    )

    # ==================================
    # CLASS SUMMARY
    # ==================================
    summary_rows = [
        ("Class", class_data["name"]),
        ("Students", class_data["students"]),
        ("Average", f'{class_data["average"]}%'),
        ("High Risk", high_count),
        ("Medium Risk", medium_count),
        ("Low Risk", low_count),
        ("Generated", datetime.now().strftime("%d/%m/%Y"))
    ]

    for row_index, (label, value) in enumerate(
        summary_rows,
        start=1
    ):

        label_cell = ws.cell(
            row=row_index,
            column=1
        )

        value_cell = ws.cell(
            row=row_index,
            column=2
        )

        if label == "Average":

            avg = class_data["average"]

            if avg >= 80:

                value_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=low_risk
                )

            elif avg >= 65:

                value_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=medium_risk
                )

            else:

                value_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=high_risk
                )

        label_cell.value = label
        value_cell.value = value

        label_cell.font = bold_font

        label_cell.alignment = center
        value_cell.alignment = center

        label_cell.border = thin_border
        value_cell.border = thin_border

    # ==================================
    # TABLE HEADER
    # ==================================
    headers = [
        "ID",
        "Student",
        "Attendance",
        "Quiz",
        "Homework",
        "Assignment",
        "Midterm",
        "Final",
        "Participation",
        "Project",
        "Behavior",
        "Average",
        "Predicted",
        "Risk"
    ]

    start_row = 11

    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=start_row,
            column=col
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # ==================================
    # STUDENT DATA
    # ==================================
    current_row = start_row + 1

    for student in students:

        average = student["average"]

        data = [
            student["id"],
            student["name"],
            student["attendance"],
            student["quiz"],
            student["homework"],
            student["assignment"],
            student["midterm"],
            student["final"],
            student["participation"],
            student["project"],
            student["behavior"],
            average,
            round(student["predicted_score"], 2),
            student["risk"]
        ]

        for col, value in enumerate(
            data,
            start=1
        ):

            cell = ws.cell(
                row=current_row,
                column=col,
                value=value
            )

            cell.border = thin_border
            cell.alignment = center

            # Average Color
            if col == 12:

                if value >= 80:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=low_risk
                    )

                elif value >= 65:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=medium_risk
                    )

                else:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=high_risk
                    )
            # Predicted Score Color
            elif col == 13:

                if value >= 80:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=low_risk
                    )

                elif value >= 65:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=medium_risk
                    )

                else:

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=high_risk
                    )
            # Risk Color
            elif col == 14:

                if value == "Low":

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=low_risk
                    )

                elif value == "Medium":

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=medium_risk
                    )

                elif value == "High":

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=high_risk
                    )

        current_row += 1

    # ==================================
    # AUTO COLUMN WIDTH
    # ==================================
    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except Exception:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 4

    # ==================================
    # FREEZE HEADER
    # ==================================
    ws.freeze_panes = "A12"

    # ==================================
    # SAVE TO DOWNLOADS
    # ==================================
    file_path = filedialog.asksaveasfilename(
        title="Save Excel Report",
        defaultextension=".xlsx",
        initialfile=f'{class_data["name"]}_report.xlsx',
        filetypes=[
            ("Excel Files", "*.xlsx")
        ]
    )

    if not file_path:
        return

    wb.save(file_path)

    print(
        f"Excel exported: {file_path}"
    )

    return str(file_path)